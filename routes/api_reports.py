import io

from flask import Blueprint, request, send_file
from flask_login import login_required
from openpyxl import Workbook

from models import Item, Location, InventoryLog, MaterialRequest, ROLE_ADMIN, ROLE_MANAGER, ROLE_STOREKEEPER
from routes.decorators import roles_required

reports_api = Blueprint('reports_api', __name__)

XLSX_MIME = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'


@reports_api.route('/inventory.xlsx', methods=['GET'])
@login_required
def export_inventory():
    locations = Location.query.order_by(Location.id).all()
    items = Item.query.order_by(Item.name).all()

    wb = Workbook()
    ws = wb.active
    ws.title = 'Inventory'
    ws.append(
        ['Item Code', 'Item Name', 'Category', 'Unit']
        + [loc.name for loc in locations]
        + ['Total Qty', 'Min. Level', 'Status']
    )

    for item in items:
        stock_by_location = {sl.location_id: sl.quantity for sl in item.stock_levels}
        total = sum(stock_by_location.values())
        status = 'Critical' if item.min_stock_level > 0 and total < item.min_stock_level else 'OK'
        ws.append(
            [item.item_code, item.name, item.category, item.unit]
            + [stock_by_location.get(loc.id, 0) for loc in locations]
            + [total, item.min_stock_level, status]
        )

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return send_file(
        buffer,
        as_attachment=True,
        download_name='inventory_export.xlsx',
        mimetype=XLSX_MIME,
    )


@reports_api.route('/logs.xlsx', methods=['GET'])
@roles_required(ROLE_ADMIN, ROLE_MANAGER, ROLE_STOREKEEPER)
def export_logs():
    action = request.args.get('action')
    reason = request.args.get('reason')

    query = InventoryLog.query
    if action:
        query = query.filter_by(action=action)
    if reason:
        query = query.filter_by(reason=reason)
    logs = query.order_by(InventoryLog.timestamp.desc()).limit(2000).all()

    wb = Workbook()
    ws = wb.active
    ws.title = 'Inventory Logs'
    ws.append([
        'Timestamp', 'User', 'Action', 'Item', 'Item Code', 'From / Location', 'To',
        'Quantity', 'Heat Number', 'Batch Number', 'Reason', 'Notes',
    ])

    for log in logs:
        ws.append([
            log.timestamp.strftime('%Y-%m-%d %H:%M') if log.timestamp else '',
            log.user.full_name if log.user else 'System',
            log.action,
            log.item.name if log.item else '',
            log.item.item_code if log.item else '',
            log.location.name if log.location else '',
            log.target_location.name if log.target_location else '',
            log.quantity if log.quantity is not None else '',
            log.heat_number or '',
            log.batch_number or '',
            log.reason or '',
            log.notes or '',
        ])

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return send_file(
        buffer,
        as_attachment=True,
        download_name='inventory_logs_export.xlsx',
        mimetype=XLSX_MIME,
    )


@reports_api.route('/requests.xlsx', methods=['GET'])
@roles_required(ROLE_ADMIN, ROLE_MANAGER, ROLE_STOREKEEPER)
def export_requests():
    status = request.args.get('status')

    query = MaterialRequest.query
    if status:
        query = query.filter_by(status=status)
    reqs = query.order_by(MaterialRequest.created_at.desc()).all()

    wb = Workbook()
    ws = wb.active
    ws.title = 'Material Requests'
    ws.append([
        'MRF #', 'Requester', 'Deliver To', 'Status', 'Created', 'Reviewed By',
        'Material', 'Item Code', 'Requested Qty', 'Received Qty', 'Unit', 'Line Status',
    ])

    for req in reqs:
        created = req.created_at.strftime('%Y-%m-%d %H:%M') if req.created_at else ''
        base_row = [
            req.id,
            req.requester.full_name if req.requester else '',
            req.target_location.name if req.target_location else '',
            req.status,
            created,
            req.reviewer.full_name if req.reviewer else '',
        ]
        if not req.lines:
            ws.append(base_row + ['', '', '', '', '', ''])
            continue
        for line in req.lines:
            material_name = line.item.name if (line.item_id and line.item) else line.new_item_name
            item_code = line.item.item_code if (line.item_id and line.item) else ''
            unit = line.item.unit if (line.item_id and line.item) else line.new_item_unit
            ws.append(base_row + [
                material_name, item_code, line.requested_quantity, line.received_quantity, unit, line.line_status,
            ])

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return send_file(
        buffer,
        as_attachment=True,
        download_name='material_requests_export.xlsx',
        mimetype=XLSX_MIME,
    )
